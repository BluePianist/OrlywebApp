from openpyxl import Workbook, load_workbook
from joueurs import Joueur
import random
from time import sleep
from statistics import stdev, mean
from math import *

wb = load_workbook('Notes_orly.xlsx', data_only=True)
ws = wb.active

nombre_equipes = 4
nombre_joueurs = 0


def get_data():
    joueurs = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,min_col=1, max_col=1):
        for cell in row:
            if cell.value != None: 
                joueurs.append(Joueur(cell.value, ceil(ws.cell(row=cell.row, column=2).value), random.randint(0,1), 0))
                
    joueurs.pop(len(joueurs)-1)
    print("liste des joueurs récupérée " + str(len(joueurs)) + " joueurs au total")
    return joueurs

def calcul_joueurs_presents(form_data, joueurs):
    joueurs_presents = []
    nombre_joueurs = 0
    moyennes = []
    # joueurs = get_data()
    # print(list(form_data))
    nombre_joueurs = len(form_data)
    print("nombre joueurs présents = " + str(nombre_joueurs))
    for data in list(form_data):
        joueurs_presents.append(joueurs[int(data)])
        moyennes.append(joueurs[int(data)].note)
    return joueurs_presents, nombre_joueurs

def calcul_equipes(joueurs_presents):

    #INITIALISATION DU NOMBRE D'EQUIPES
    moyennes_equipes = []
    moyennes = []
    equipes = []
    equipes.clear
    for equipe in range(nombre_equipes):
        equipes.append([])
        moyennes_equipes.append(0)
        
    # get_data()
    

    nombre_joueurs = len(joueurs_presents)
    for joueur in joueurs_presents:
            moyennes.append(joueur.note)
    moyenne_totale = mean(moyennes)

    print("joueurs présents = ", joueurs_presents)
    print(nombre_joueurs)
    equipe_max = nombre_joueurs%nombre_equipes # NOMBRE D'EQUIPE AVEC UN REMPLACANT
    joueurs_par_equipe = nombre_joueurs/nombre_equipes
    print("\njoueurs par equipes = ", joueurs_par_equipe)
    # print(equipe_max)
    
        
    i = 0
    for joueur in joueurs_presents:
        while True:
            alea = random.randint(1, nombre_equipes)
            # print(joueurs_par_equipe, alea, len(equipes[alea-1]), i, joueur.name)
            # sleep(0.1)
            
            if (len(equipes[alea-1]) < joueurs_par_equipe): #SI L'EQUIPE NE CONTIENT PAS DEJA PLUS DE x JOUEURS
                if(i > equipe_max and len(equipes[alea-1]) >= int(joueurs_par_equipe)): #SI ON A ATTEINT LE NOMBRE MAX D'EQUIPES AVEC REMPLACANT ET QUE CETTE EQUIPE EST PLEINE
                    continue
                else:
                    equipes[alea-1].append(joueur)
                    joueur.equipe = alea
                    if(len(equipes[alea-1]) == int(joueurs_par_equipe)+1 and equipe_max != 0): 
                        i += 1.1
                        
                    moyennes_equipes[alea-1] = 0
                    for joueur_selected in equipes[alea-1]:
                        moyennes_equipes[alea-1] += joueur_selected.note
                    moyennes_equipes[alea-1] = float("{:.1f}".format(moyennes_equipes[alea-1]/len(equipes[alea-1])))
                    # print("moyenne equipe " + str(alea) + "=" + str(moyennes_equipes[alea-1]))
                    break
            else :
                continue
            
    print("moyenne totale = "+ str(moyenne_totale) +" ecart type = ", str(stdev(moyennes)))
    for indx, equipe in enumerate(equipes):
        # for i in equipe:
        #     print(i.name, i.equipe, i.note)
        print("equipe " + str(indx+1) + " moyenne = ", moyennes_equipes[indx])
    return equipes, moyennes_equipes

# calcul_equipes()
# get_data()
